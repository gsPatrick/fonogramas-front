# usuario/services/upload_service.py
from shared.processador import processar_arquivo_fonogramas
from shared.fonograma_service import salvar_fonogramas_do_dataframe
from shared.validador import validar_isrc, validar_cpf, validar_cnpj, validar_duracao
from models import db, Fonograma
from datetime import datetime
import pandas as pd
import tempfile
import os

def validar_arquivo(arquivo):
    """Valida arquivo e retorna preview para usuário"""
    # Salvar temporariamente
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(arquivo.filename)[1])
    arquivo.save(temp.name)
    
    # Processar
    resultado = processar_arquivo_fonogramas(temp.name)
    
    # Enriquecer com validações extras e pré-check de duplicidade
    validos = []
    com_erro = []
    
    for item in resultado['dados']:
        erros = item.get('erros', [])
        avisos = []
        
        # Verificar duplicidade
        if Fonograma.query.filter_by(isrc=item.get('isrc')).first():
            avisos.append('ISRC já existe (será atualizado)')
            
        if erros:
            item['erros'] = erros
            item['avisos'] = avisos
            com_erro.append(item)
        else:
            item['avisos'] = avisos
            validos.append(item)
            
    return {
        'total': len(resultado['dados']),
        'validos': validos,
        'com_erro': com_erro,
        'arquivo_temp': temp.name
    }

def salvar_fonogramas(resultado_validacao, usuario, salvar_apenas_validos=False):
    """Salva fonogramas da validação usando dados diretamente da sessão"""
    try:
        # Tentar usar dados validados diretamente (mais confiável)
        dados_validados = resultado_validacao.get('dados_validados', [])
        
        if dados_validados:
            df = pd.DataFrame(dados_validados)
            # Remover colunas internas
            cols_to_drop = ['erros', 'avisos', 'linha']
            df = df.drop(columns=[c for c in cols_to_drop if c in df.columns], errors='ignore')
        else:
            # Fallback: tentar ler do arquivo
            arquivo_temp = resultado_validacao.get('arquivo_temp')
            
            if not arquivo_temp or not os.path.exists(arquivo_temp):
                return {'sucesso': False, 'erro': 'Nenhum dado disponível para salvar'}
            
            df = pd.read_excel(arquivo_temp)
        
        # Importar app para obter contexto
        from app import app
        
        # Chamar com parâmetros corretos - incluindo user_id para ownership
        resultado = salvar_fonogramas_do_dataframe(df, app.app_context(), salvar_apenas_validos, user_id=usuario.id)
        
        # Adicionar flag de sucesso
        resultado['sucesso'] = True
        return resultado
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'sucesso': False, 'erro': str(e)}

def gerar_template():
    """Gera template Excel para download"""
    # ... Lógica de geração de template ou retornar estático
    # Para simplificar, vamos criar um df vazio com colunas
    colunas = [
        'isrc', 'titulo', 'duracao', 'ano_lanc', 'genero', 'versao',
        'autores', 'interpretes', 'musicos', 'produtor'
    ]
    df = pd.DataFrame(columns=colunas)
    arquivo = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    df.to_excel(arquivo.name, index=False)
    return arquivo.name

def validar_campo_individual(campo, valor):
    """Valida um campo individualmente"""
    if campo == 'isrc':
        return {'valido': validar_isrc(valor)}
    elif campo == 'cpf':
        return {'valido': validar_cpf(valor)}
    elif campo == 'cnpj':
        return {'valido': validar_cnpj(valor)}
    elif campo == 'duracao':
        return {'valido': validar_duracao(valor)}
    return {'valido': True}

def processar_para_frontend(arquivo):
    """
    Processa arquivo e retorna formato esperado pelo script.js antigo.
    Gera Excel dos dados válidos no formato template_fonograma_final.
    """
    # Mapeamento de colunas internas -> Template ECAD
    MAPEAMENTO_COLUNAS_EXPORT = {
        'isrc': 'ISRC *',
        'titulo': 'Título *',
        'duracao': 'Duração *',
        'ano_lanc': 'Ano Lanc. *',
        'genero': 'Gênero *',
        'titulo_obra': 'Título Obra *',
        'autores': 'Autores * (Nome|CPF|Função|%)',
        'interpretes': 'Intérpretes (Nome|Doc|Cat|%|Assoc)',
        'prod_nome': 'Produtor Nome *',
        'prod_doc': 'Produtor Doc *',
        'prod_perc': 'Produtor % *',
        'versao': 'Versão',
        'idioma': 'Idioma',
        'ano_grav': 'Ano Grav.',
        'cod_interno': 'Cód. Interno',
        'cod_obra': 'Cód. Obra',
        'editoras': 'Editoras (Nome|CNPJ|%)',
        'musicos': 'Músicos (Nome|CPF|Instr|Tipo|%)',
        'prod_fantasia': 'Prod. Fantasia',
        'prod_assoc': 'Prod. Assoc.',
        'tipo_lanc': 'Tipo Lanç.',
        'album': 'Álbum',
        'faixa': 'Faixa',
        'formato': 'Formato',
        'situacao': 'Situação',
        'territorio': 'Território',
    }
    
    # 1. Validar
    resultado = validar_arquivo(arquivo)
    
    # 2. Gerar Excel de saída (apenas válidos para converter)
    arquivo_excel = None
    arquivo_save = None
    if resultado['validos']:
        # DataFrame original com colunas lowercase (para save)
        df_original = pd.DataFrame(resultado['validos'])
        # Remover colunas internas de controle se existirem
        cols_to_drop = ['erros', 'avisos', 'linha']
        df_original = df_original.drop(columns=[c for c in cols_to_drop if c in df_original.columns], errors='ignore')
        
        # DataFrame para template (colunas renomeadas para ECAD)
        df_template = df_original.copy()
        df_template = df_template.rename(columns=MAPEAMENTO_COLUNAS_EXPORT)
        
        # Ordenar colunas na ordem do template
        ordem_colunas = list(MAPEAMENTO_COLUNAS_EXPORT.values())
        colunas_existentes = [c for c in ordem_colunas if c in df_template.columns]
        colunas_extras = [c for c in df_template.columns if c not in ordem_colunas]
        df_template = df_template[colunas_existentes + colunas_extras]
        
        # Gerar arquivo - usar delete=False e fechar antes de mover
        import shutil
        output_dir = os.path.join(os.getcwd(), 'outputs')
        os.makedirs(output_dir, exist_ok=True)
        
        # 1. Gerar Excel no formato TEMPLATE para download (com colunas ISRC *, Título *, etc.)
        temp_excel = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_path = temp_excel.name
        temp_excel.close()  # IMPORTANTE: Fechar antes de escrever para evitar lock
        
        # Escrever DataFrame com estilos idênticos ao export_service
        _gerar_excel_estilizado(df_template, temp_path)
        arquivo_excel = os.path.basename(temp_path)
        
        # Mover para pasta de outputs
        final_path = os.path.join(output_dir, arquivo_excel)
        shutil.copy2(temp_path, final_path)
        try:
            os.remove(temp_path)
        except:
            pass  # Se não conseguir remover, tudo bem - é temporário
        
        # 2. Gerar Excel com dados ORIGINAIS para save no banco (colunas lowercase)
        temp_save = tempfile.NamedTemporaryFile(suffix='_save.xlsx', delete=False)
        temp_save_path = temp_save.name
        temp_save.close()
        df_original.to_excel(temp_save_path, index=False)  # DataFrame original com colunas lowercase
        
        # Mover para pasta de outputs
        arquivo_save = os.path.basename(temp_save_path)
        final_save_path = os.path.join(output_dir, arquivo_save)
        shutil.copy2(temp_save_path, final_save_path)
        try:
            os.remove(temp_save_path)
        except:
            pass
    
    # 3. Formatar Erros (Flatten)
    erros_flat = []
    total_erros = 0
    for i, item in enumerate(resultado['com_erro']):
        linha_num = i + 1 + len(resultado['validos']) # Aproximação ou usar índice real se tiver
        # O processador original talvez retornasse linha. Vamos usar um contador.
        # Na verdade, o validar_arquivo retorna itens. Vamos assumir que 'linha' está nos dados ou usar contador.
        linha_real = item.get('linha', i + 1)
        
        for erro_msg in item.get('erros', []):
            total_erros += 1
            # Tentar extrair campo se a mensagem for "Campo X: erro"
            campo = '-'
            msg = erro_msg
            if ':' in erro_msg:
                parts = erro_msg.split(':', 1)
                campo = parts[0].strip()
                msg = parts[1].strip()
            
            erros_flat.append({
                'linha': linha_real,
                'campo': campo,
                'valor': str(item.get(campo.lower(), '-')),
                'erro': msg
            })

    # Obter caminhos COMPLETOS dos arquivos gerados
    arquivo_excel_path = None
    arquivo_save_path = None
    if arquivo_excel:
        arquivo_excel_path = os.path.join(os.getcwd(), 'outputs', arquivo_excel)
    if arquivo_save:
        arquivo_save_path = os.path.join(os.getcwd(), 'outputs', arquivo_save)

    return {
        'total_linhas': resultado['total'],
        'linhas_validas': len(resultado['validos']),
        'linhas_com_erro': len(resultado['com_erro']),
        'total_erros': total_erros,
        'erros': erros_flat,
        'arquivo_excel': arquivo_excel,  # Nome para download URL (formato template)
        'arquivo_temp': arquivo_save_path or arquivo_excel_path,  # Caminho do arquivo (fallback)
        'dados_validados': resultado['validos']  # Dados originais para salvar no banco
    }


def _gerar_excel_estilizado(df, filepath):
    """
    Gera Excel com estilos SBACEM idênticos ao export_service.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fonogramas"
    
    # Headers do DataFrame
    headers = list(df.columns)
    
    # === CORES (EXATAS do template) ===
    cor_identificacao = PatternFill(start_color="22164C", end_color="22164C", fill_type="solid")
    cor_obra = PatternFill(start_color="3D2E6B", end_color="3D2E6B", fill_type="solid")
    cor_autores = PatternFill(start_color="EF234D", end_color="EF234D", fill_type="solid")
    cor_produtor = PatternFill(start_color="E1C8B0", end_color="E1C8B0", fill_type="solid")
    cor_opcional = PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid")
    cor_dados = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    fonte_branca = Font(color="FFFFFF", bold=True)
    fonte_escura = Font(color="22164C", bold=True)
    fonte_dados = Font(color="000000")
    
    alinhamento_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alinhamento_dados = Alignment(vertical="center", wrap_text=True)
    
    borda = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Mapa de cores por header
    def get_header_style(header):
        header_lower = header.lower()
        if 'isrc' in header_lower or 'título *' == header_lower or 'duração' in header_lower or 'ano lanc' in header_lower or 'gênero' in header_lower:
            return cor_identificacao, fonte_branca
        elif 'título obra' in header_lower:
            return cor_obra, fonte_branca
        elif 'autores' in header_lower or 'intérpretes' in header_lower:
            return cor_autores, fonte_branca
        elif 'produtor' in header_lower:
            return cor_produtor, fonte_escura
        else:
            return cor_opcional, fonte_branca
    
    # Escrever headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cor, fonte = get_header_style(header)
        cell.fill = cor
        cell.font = fonte
        cell.alignment = alinhamento_header
        cell.border = borda
    
    # Larguras de colunas
    larguras = {'A': 14, 'B': 25, 'C': 10, 'D': 10, 'E': 12, 'F': 25, 'G': 50, 'H': 45, 'I': 25, 'J': 16, 'K': 10}
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width
    
    # Altura do header
    ws.row_dimensions[1].height = 40
    
    # Escrever dados
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = cor_dados
            cell.font = fonte_dados
            cell.alignment = alinhamento_dados
            cell.border = borda
        ws.row_dimensions[row_idx].height = 25
    
    # Congelar header
    ws.freeze_panes = 'A2'
    
    # Salvar
    wb.save(filepath)
    wb.close()

