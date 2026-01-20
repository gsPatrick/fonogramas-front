# usuario/services/export_service.py
"""
Exportação de fonogramas para Excel seguindo o visual EXATO do template_fonograma_final.xlsx
Cria arquivo do zero para evitar corrupção ao modificar template
"""
import os
import tempfile
from models import Fonograma


def exportar_fonogramas(usuario, fonograma_ids=None):
    """
    Exporta fonogramas para Excel com estilo idêntico ao template.
    Cria arquivo do zero para evitar erro de corrupção do Excel.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    query = Fonograma.query.filter_by(user_id=usuario.id)
        
    if fonograma_ids:
        query = query.filter(Fonograma.id.in_(fonograma_ids))
        
    fonogramas = query.order_by(Fonograma.created_at.desc()).all()
    
    # Criar workbook novo (evita problemas de corrupção ao copiar template)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fonogramas"
    
    # === DEFINIR HEADERS ===
    headers = [
        'ISRC *', 'Título *', 'Duração *', 'Ano Lanc. *', 'Gênero *', 'Título Obra *',
        'Autores * (Nome|CPF|Função|%)', 'Intérpretes (Nome|Doc|Cat|%|Assoc)',
        'Produtor Nome *', 'Produtor Doc *', 'Produtor % *',
        'Versão', 'Idioma', 'Ano Grav.', 'Cód. Interno', 'Cód. Obra',
        'Editoras (Nome|CNPJ|%)', 'Músicos (Nome|CPF|Instr|Tipo|%)',
        'Prod. Fantasia', 'Prod. Assoc.', 'Tipo Lanç.', 'Álbum', 'Faixa',
        'Formato', 'Situação', 'Território'
    ]
    
    # === CORES (EXATAS do template) ===
    # Identificação: 22164C (roxo escuro)
    cor_identificacao = PatternFill(start_color="22164C", end_color="22164C", fill_type="solid")
    # Obra: 3D2E6B (roxo médio)
    cor_obra = PatternFill(start_color="3D2E6B", end_color="3D2E6B", fill_type="solid")
    # Autores/Intérpretes: EF234D (vermelho SBACEM)
    cor_autores = PatternFill(start_color="EF234D", end_color="EF234D", fill_type="solid")
    # Produtor: E1C8B0 (bege)
    cor_produtor = PatternFill(start_color="E1C8B0", end_color="E1C8B0", fill_type="solid")
    # Opcionais: 4A4A4A (cinza)
    cor_opcional = PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid")
    # Dados: FFF3CD (amarelo claro)
    cor_dados = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # Fontes
    fonte_branca = Font(color="FFFFFF", bold=True)
    fonte_escura = Font(color="22164C", bold=True)
    fonte_dados = Font(color="000000")
    
    # Alinhamentos
    alinhamento_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alinhamento_dados = Alignment(vertical="center", wrap_text=True)
    
    # Borda
    borda = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Mapa de cores por coluna (index 0-based)
    cores_colunas = [
        (cor_identificacao, fonte_branca),  # 1 ISRC
        (cor_identificacao, fonte_branca),  # 2 Título
        (cor_identificacao, fonte_branca),  # 3 Duração
        (cor_identificacao, fonte_branca),  # 4 Ano Lanc
        (cor_identificacao, fonte_branca),  # 5 Gênero
        (cor_obra, fonte_branca),           # 6 Título Obra
        (cor_autores, fonte_branca),        # 7 Autores
        (cor_autores, fonte_branca),        # 8 Intérpretes
        (cor_produtor, fonte_escura),       # 9 Produtor Nome
        (cor_produtor, fonte_escura),       # 10 Produtor Doc
        (cor_produtor, fonte_escura),       # 11 Produtor %
        (cor_opcional, fonte_branca),       # 12 Versão
        (cor_opcional, fonte_branca),       # 13 Idioma
        (cor_opcional, fonte_branca),       # 14 Ano Grav
        (cor_opcional, fonte_branca),       # 15 Cód Interno
        (cor_opcional, fonte_branca),       # 16 Cód Obra
        (cor_opcional, fonte_branca),       # 17 Editoras
        (cor_opcional, fonte_branca),       # 18 Músicos
        (cor_opcional, fonte_branca),       # 19 Prod Fantasia
        (cor_opcional, fonte_branca),       # 20 Prod Assoc
        (cor_opcional, fonte_branca),       # 21 Tipo Lanç
        (cor_opcional, fonte_branca),       # 22 Álbum
        (cor_opcional, fonte_branca),       # 23 Faixa
        (cor_opcional, fonte_branca),       # 24 Formato
        (cor_opcional, fonte_branca),       # 25 Situação
        (cor_opcional, fonte_branca),       # 26 Território
    ]
    
    # === ESCREVER HEADERS ===
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cor, fonte = cores_colunas[col_idx - 1]
        cell.fill = cor
        cell.font = fonte
        cell.alignment = alinhamento_header
        cell.border = borda
    
    # === LARGURAS DAS COLUNAS (idêntico ao template) ===
    larguras = {
        'A': 14, 'B': 25, 'C': 10, 'D': 10, 'E': 12, 'F': 25,
        'G': 50, 'H': 45, 'I': 25, 'J': 16, 'K': 10,
        'L': 12, 'M': 8, 'N': 10, 'O': 12, 'P': 12,
        'Q': 30, 'R': 35, 'S': 15, 'T': 12, 'U': 12,
        'V': 20, 'W': 8, 'X': 10, 'Y': 10, 'Z': 12
    }
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width
    
    # === ALTURA DO HEADER ===
    ws.row_dimensions[1].height = 40
    
    # === INSERIR DADOS ===
    for row_idx, f in enumerate(fonogramas, 2):
        # Formatar relações no formato pipe-delimited
        autores_str = '; '.join([
            f"{a.nome}|{a.cpf}|{a.funcao}|{a.percentual}" 
            for a in f.autores_list
        ]) if f.autores_list else ''
        
        interpretes_str = '; '.join([
            f"{i.nome}|{i.doc}|{i.categoria}|{i.percentual}|{i.associacao or ''}"
            for i in f.interpretes_list
        ]) if f.interpretes_list else ''
        
        editoras_str = '; '.join([
            f"{e.nome}|{e.cnpj}|{e.percentual}"
            for e in f.editoras_list
        ]) if f.editoras_list else ''
        
        musicos_str = '; '.join([
            f"{m.nome}|{m.cpf}|{m.instrumento}|{m.tipo}|{m.percentual}"
            for m in f.musicos_list
        ]) if f.musicos_list else ''
        
        # Dados na ordem do template
        row_data = [
            f.isrc,
            f.titulo,
            f.duracao,
            f.ano_lanc,
            f.genero,
            f.titulo_obra,
            autores_str,
            interpretes_str,
            f.prod_nome,
            f.prod_doc,
            f.prod_perc,
            f.versao,
            f.idioma,
            f.ano_grav,
            f.cod_interno,
            f.cod_obra,
            editoras_str,
            musicos_str,
            f.prod_fantasia,
            f.prod_assoc,
            f.tipo_lanc,
            f.album,
            f.faixa,
            f.formato,
            f.situacao,
            f.territorio,
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = cor_dados
            cell.font = fonte_dados
            cell.alignment = alinhamento_dados
            cell.border = borda
        
        # Altura da linha de dados
        ws.row_dimensions[row_idx].height = 25
    
    # === CONGELAR HEADER ===
    ws.freeze_panes = 'A2'
    
    # === SALVAR ARQUIVO ===
    arquivo = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    arquivo.close()  # Fechar o handle antes de salvar
    wb.save(arquivo.name)
    wb.close()  # Fechar o workbook também
    
    return arquivo.name
