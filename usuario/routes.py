# usuario/routes.py
import os
from flask import Blueprint, render_template, request, jsonify, send_file, flash, redirect, url_for, session
from flask_login import current_user
from shared.decorators import usuario_required
from models import db, Fonograma
from usuario.services import fonograma_service, upload_service, export_service
from . import usuario_bp

# ==================== DASHBOARD ====================

@usuario_bp.route('/')
@usuario_required
def dashboard():
    """Dashboard do usuário"""
    stats = fonograma_service.obter_estatisticas_usuario(current_user)
    recentes = fonograma_service.obter_fonogramas_recentes(current_user, limit=10)
    return render_template('usuario/dashboard.html', stats=stats, recentes=recentes)

# ==================== FONOGRAMAS ====================

@usuario_bp.route('/fonogramas')
@usuario_required
def listar_fonogramas():
    """Lista fonogramas do usuário com filtros combinados"""
    from datetime import datetime
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    busca = request.args.get('busca', '')
    genero = request.args.get('genero')
    status = request.args.get('status')
    prod_assoc = request.args.get('prod_assoc')
    data_inicio_str = request.args.get('data_inicio')
    data_fim_str = request.args.get('data_fim')
    
    # Parsear datas
    data_inicio = None
    data_fim = None
    if data_inicio_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d')
        except ValueError:
            pass
    if data_fim_str:
        try:
            data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d')
        except ValueError:
            pass
    
    fonogramas = fonograma_service.listar_fonogramas(
        usuario=current_user,
        page=page,
        per_page=per_page,
        busca=busca,
        genero=genero,
        status=status,
        prod_assoc=prod_assoc,
        data_inicio=data_inicio,
        data_fim=data_fim
    )
    
    # Passar filtros atuais para o template (manter estado dos filtros)
    filtros = {
        'busca': busca,
        'genero': genero,
        'status': status,
        'prod_assoc': prod_assoc,
        'data_inicio': data_inicio_str or '',
        'data_fim': data_fim_str or ''
    }
    
    return render_template('usuario/fonogramas/lista.html', fonogramas=fonogramas, filtros=filtros)

@usuario_bp.route('/fonogramas/novo', methods=['GET', 'POST'])
@usuario_required
def novo_fonograma():
    """Criar novo fonograma manualmente"""
    if request.method == 'POST':
        dados = request.form.to_dict()
        resultado = fonograma_service.criar_fonograma(dados, current_user)
        
        if resultado['sucesso']:
            flash('Fonograma criado com sucesso!', 'success')
            return redirect(url_for('usuario.detalhes_fonograma', fonograma_id=resultado['fonograma_id']))
        else:
            flash(f'Erro: {resultado["erro"]}', 'danger')
            return render_template('usuario/fonogramas/novo.html', dados=dados, erros=resultado.get('erros', []))
    
    return render_template('usuario/fonogramas/novo.html')

@usuario_bp.route('/fonogramas/<int:fonograma_id>')
@usuario_required
def detalhes_fonograma(fonograma_id):
    """Detalhes de um fonograma"""
    fonograma = fonograma_service.obter_fonograma(fonograma_id, current_user)
    if not fonograma:
        flash('Fonograma não encontrado.', 'danger')
        return redirect(url_for('usuario.listar_fonogramas'))
    
    # Status simplificado (sem detalhes técnicos)
    status_display = fonograma_service.obter_status_simplificado(fonograma)
    
    return render_template('usuario/fonogramas/detalhes.html', 
                         fonograma=fonograma, 
                         status_display=status_display)

@usuario_bp.route('/fonogramas/<int:fonograma_id>/editar', methods=['GET', 'POST'])
@usuario_required
def editar_fonograma(fonograma_id):
    """Editar fonograma"""
    fonograma = fonograma_service.obter_fonograma(fonograma_id, current_user)
    if not fonograma:
        flash('Fonograma não encontrado.', 'danger')
        return redirect(url_for('usuario.listar_fonogramas'))
    
    # Não permitir edição se já foi enviado/aceito
    if fonograma.status_ecad in ['ENVIADO', 'ACEITO']:
        flash('Fonograma já enviado ao ECAD não pode ser editado.', 'warning')
        return redirect(url_for('usuario.detalhes_fonograma', fonograma_id=fonograma_id))
    
    if request.method == 'POST':
        dados = request.form.to_dict()
        resultado = fonograma_service.atualizar_fonograma(fonograma, dados, current_user)
        
        if resultado['sucesso']:
            flash('Fonograma atualizado com sucesso!', 'success')
            return redirect(url_for('usuario.detalhes_fonograma', fonograma_id=fonograma_id))
        else:
            flash(f'Erro: {resultado["erro"]}', 'danger')
    
    return render_template('usuario/fonogramas/editar.html', fonograma=fonograma)

@usuario_bp.route('/fonogramas/<int:fonograma_id>/excluir', methods=['POST'])
@usuario_required
def excluir_fonograma(fonograma_id):
    """Excluir fonograma"""
    fonograma = fonograma_service.obter_fonograma(fonograma_id, current_user)
    if not fonograma:
        flash('Fonograma não encontrado.', 'danger')
        return redirect(url_for('usuario.listar_fonogramas'))
    
    # Não permitir exclusão se já foi enviado
    if fonograma.status_ecad in ['ENVIADO', 'ACEITO']:
        flash('Fonograma já enviado ao ECAD não pode ser excluído.', 'warning')
        return redirect(url_for('usuario.detalhes_fonograma', fonograma_id=fonograma_id))
    
    resultado = fonograma_service.excluir_fonograma(fonograma, current_user)
    
    if resultado['sucesso']:
        flash('Fonograma excluído com sucesso!', 'success')
    else:
        flash(f'Erro: {resultado["erro"]}', 'danger')
    
    return redirect(url_for('usuario.listar_fonogramas'))

@usuario_bp.route('/fonogramas/excluir-massa', methods=['POST'])
@usuario_required
def excluir_fonogramas_massa():
    """Excluir múltiplos fonogramas de uma vez"""
    ids_str = request.form.get('fonograma_ids', '')
    
    if not ids_str:
        flash('Nenhum fonograma selecionado.', 'warning')
        return redirect(url_for('usuario.listar_fonogramas'))
    
    # Converter string de IDs para lista
    try:
        ids = [int(id.strip()) for id in ids_str.split(',') if id.strip()]
    except ValueError:
        flash('IDs inválidos.', 'danger')
        return redirect(url_for('usuario.listar_fonogramas'))
    
    excluidos = 0
    erros = 0
    
    for fono_id in ids:
        fonograma = fonograma_service.obter_fonograma(fono_id, current_user)
        if not fonograma:
            erros += 1
            continue
        
        # Não permitir exclusão se já foi enviado
        if fonograma.status_ecad in ['ENVIADO', 'ACEITO']:
            erros += 1
            continue
        
        resultado = fonograma_service.excluir_fonograma(fonograma, current_user)
        if resultado['sucesso']:
            excluidos += 1
        else:
            erros += 1
    
    if excluidos > 0 and erros == 0:
        flash(f'{excluidos} fonograma(s) excluído(s) com sucesso!', 'success')
    elif excluidos > 0 and erros > 0:
        flash(f'{excluidos} excluído(s), {erros} não puderam ser excluídos (enviados ou não encontrados).', 'warning')
    else:
        flash('Nenhum fonograma foi excluído.', 'danger')
    
    return redirect(url_for('usuario.listar_fonogramas'))

@usuario_bp.route('/fonogramas/editar-massa', methods=['POST'])
@usuario_required
def editar_fonogramas_massa():
    """Editar múltiplos fonogramas de uma vez"""
    from models import db
    
    ids_str = request.form.get('fonograma_ids', '')
    campo = request.form.get('campo', '').strip()
    valor = request.form.get('valor', '').strip()
    
    # Validações
    if not ids_str:
        flash('Nenhum fonograma selecionado.', 'warning')
        return redirect(url_for('usuario.listar_fonogramas'))
    
    if not campo:
        flash('Selecione um campo para editar.', 'warning')
        return redirect(url_for('usuario.listar_fonogramas'))
    
    # Converter string de IDs para lista
    try:
        ids = [int(id.strip()) for id in ids_str.split(',') if id.strip()]
    except ValueError:
        flash('IDs inválidos.', 'danger')
        return redirect(url_for('usuario.listar_fonogramas'))
    
    if not ids:
        flash('Nenhum fonograma selecionado.', 'warning')
        return redirect(url_for('usuario.listar_fonogramas'))
    
    # Campos que podem ser editados em lote
    campos_permitidos = ['genero', 'prod_assoc', 'assoc_gestao', 'situacao', 'versao', 'idioma']
    
    if campo not in campos_permitidos:
        flash(f'Campo "{campo}" não permitido para edição em lote.', 'danger')
        return redirect(url_for('usuario.listar_fonogramas'))
    
    editados = 0
    erros = 0
    
    try:
        for fono_id in ids:
            fonograma = fonograma_service.obter_fonograma(fono_id, current_user)
            if not fonograma:
                erros += 1
                continue
            
            # Não permitir edição se já foi enviado/aceito
            if fonograma.status_ecad in ['ENVIADO', 'ACEITO']:
                erros += 1
                continue
            
            setattr(fonograma, campo, valor if valor else None)
            editados += 1
        
        # Commit único no final para melhor performance
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao editar fonogramas: {str(e)}', 'danger')
        return redirect(url_for('usuario.listar_fonogramas'))
    
    if editados > 0 and erros == 0:
        flash(f'{editados} fonograma(s) atualizado(s) com sucesso!', 'success')
    elif editados > 0 and erros > 0:
        flash(f'{editados} atualizado(s), {erros} não puderam ser editados.', 'warning')
    else:
        flash('Nenhum fonograma foi editado.', 'danger')
    
    return redirect(url_for('usuario.listar_fonogramas'))

# ==================== UPLOAD ====================

@usuario_bp.route('/upload', methods=['GET', 'POST'])
@usuario_required
def upload_arquivo():
    """Upload de arquivo para validação"""
    if request.method == 'POST':
        if 'arquivo' not in request.files and 'file' not in request.files:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'erro': 'Nenhum arquivo enviado'}), 400
            flash('Nenhum arquivo enviado.', 'danger')
            return redirect(request.url)
        
        arquivo = request.files.get('arquivo') or request.files.get('file')
        
        # Detecção de AJAX/Fetch do script.js
        # O script.js original envia via fetch. Vamos detectar se espera JSON.
        # Mas para garantir, vamos ver se o campo 'file' (usado pelo script.js) está presente
        is_ajax = 'file' in request.files or request.path.endswith('/upload') # script.js usa /upload se ajustarmos
        
        if is_ajax:
            # Modo API para o Conversor (script.js)
            try:
                if not arquivo.filename.lower().endswith('.csv'):
                     return jsonify({'erro': 'Apenas arquivos CSV são aceitos para conversão.'}), 400

                resultado = upload_service.processar_para_frontend(arquivo)
                
                # IMPORTANTE: Salvar dados VALIDADOS na sessão para o botão "Salvar no Dashboard"
                # Agora salvamos os dados brutos validados diretamente, não o caminho do arquivo
                session['upload_resultado'] = {
                    'arquivo_temp': resultado.get('arquivo_temp'),
                    'dados_validados': resultado.get('dados_validados', [])  # Lista de dicts com dados originais
                }
                session.modified = True  # Forçar persistência da sessão
                
                return jsonify(resultado)
            except Exception as e:
                import traceback
                traceback.print_exc()
                return jsonify({'erro': str(e)}), 500

        # Modo Clássico (Formulário Padrão - Mantendo compatibilidade se necessário)
        resultado = upload_service.validar_arquivo(arquivo)
        session['upload_resultado'] = resultado
        return render_template('usuario/upload/resultado.html', resultado=resultado)
    
    return render_template('usuario/upload/form.html')

@usuario_bp.route('/download/<filename>')
@usuario_required
def download_arquivo(filename):
    """Download de arquivos gerados (Excel do conversor)"""
    try:
        # Assumindo que os arquivos estão na pasta 'outputs' na raiz do app
        output_dir = os.path.join(os.getcwd(), 'outputs')
        filepath = os.path.join(output_dir, filename)
        
        # Garantir que o nome do arquivo tenha extensão .xlsx
        download_name = filename if filename.endswith('.xlsx') else f'{filename}.xlsx'
        
        return send_file(
            filepath, 
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Erro ao baixar arquivo: {str(e)}', 'danger')
        return redirect(url_for('usuario.upload_arquivo'))

@usuario_bp.route('/upload/salvar', methods=['POST'])
@usuario_required
def salvar_upload():
    """Salvar fonogramas validados"""
    resultado_validacao = session.get('upload_resultado')
    
    if not resultado_validacao:
        flash('Sessão expirada. Faça o upload novamente.', 'warning')
        return redirect(url_for('usuario.upload_arquivo'))
    
    # Filtrar apenas linhas válidas
    salvar_apenas_validos = request.form.get('salvar_apenas_validos') == 'on'
    
    resultado = upload_service.salvar_fonogramas(
        resultado_validacao,
        current_user,
        salvar_apenas_validos=salvar_apenas_validos
    )
    
    if resultado['sucesso']:
        flash(f'Salvos: {resultado["salvos"]}, Atualizados: {resultado["atualizados"]}', 'success')
        session.pop('upload_resultado', None)
        return redirect(url_for('usuario.listar_fonogramas'))
    else:
        flash(f'Erro: {resultado["erro"]}', 'danger')
        return redirect(url_for('usuario.upload_arquivo'))

@usuario_bp.route('/upload/template')
@usuario_required
def download_template():
    """Download do template de importação (CSV ou Excel)"""
    import os
    import io
    from flask import current_app, make_response
    
    # Verificar formato solicitado (csv ou xlsx)
    formato = request.args.get('formato', 'csv').lower()
    
    if formato == 'xlsx':
        # Gerar template Excel estilizado
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template Fonogramas"
        
        # Headers
        headers = [
            'isrc', 'titulo', 'versao', 'duracao', 'ano_grav', 'ano_lanc', 
            'idioma', 'genero', 'titulo_obra', 'autores', 'interpretes', 
            'prod_nome', 'prod_doc', 'prod_perc', 'prod_assoc'
        ]
        
        # Estilos
        header_fill = PatternFill(start_color="22164C", end_color="22164C", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Escrever headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border
        
        # Exemplo de dados
        exemplo = [
            'BRUM71601234', 'Música Exemplo', 'original', '03:45', 2023, 2024,
            'PT', 'Pop', 'Obra Musical', 'João Silva|11144477735|COMPOSITOR|100',
            'Carlos Oliveira|11144477735|PRINCIPAL|100|ABRAMUS',
            'Produtora XYZ', '11222333000181', 100, 'ABRAMUS'
        ]
        
        data_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        for col, valor in enumerate(exemplo, 1):
            cell = ws.cell(row=2, column=col, value=valor)
            cell.fill = data_fill
            cell.border = border
        
        # Ajustar larguras
        larguras = {'A': 14, 'B': 25, 'C': 12, 'D': 10, 'E': 10, 'F': 10, 
                   'G': 8, 'H': 12, 'I': 25, 'J': 45, 'K': 45, 
                   'L': 22, 'M': 16, 'N': 10, 'O': 12}
        for col, width in larguras.items():
            ws.column_dimensions[col].width = width
        
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=template_fonogramas.xlsx'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return response
    
    else:
        # Template CSV
        root_path = current_app.root_path
        arquivo_exemplo = os.path.join(root_path, 'exemplo.csv')
        
        if os.path.exists(arquivo_exemplo):
            with open(arquivo_exemplo, 'rb') as f:
                content = f.read()
            
            response = make_response(content)
            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
            response.headers['Content-Disposition'] = 'attachment; filename=exemplo_fonogramas.csv'
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response
        else:
            # Fallback: gerar template em memória
            csv_content = """isrc,titulo,versao,duracao,ano_grav,ano_lanc,idioma,genero,titulo_obra,autores,interpretes,prod_nome,prod_doc
BRUM71601234,Música Exemplo,original,03:45,2023,2024,PT,Pop,Obra Musical,João Silva|11144477735|COMPOSITOR|100,Carlos Oliveira|11144477735|PRINCIPAL|100|ABRAMUS,Produtora XYZ,11222333000181"""
            
            buffer = io.BytesIO()
            buffer.write(csv_content.encode('utf-8'))
            buffer.seek(0)
            
            return send_file(buffer, as_attachment=True, download_name='exemplo_fonogramas.csv', mimetype='text/csv')

# ==================== EXPORTAÇÃO ====================

@usuario_bp.route('/exportar', methods=['GET', 'POST'])
@usuario_required
def exportar():
    """Exportar fonogramas para Excel"""
    if request.method == 'POST':
        fonograma_ids = request.form.getlist('fonograma_ids')
        
        if not fonograma_ids:
            # Exportar todos os fonogramas do usuário
            fonograma_ids = None
        
        arquivo = export_service.exportar_fonogramas(
            usuario=current_user,
            fonograma_ids=fonograma_ids
        )
        
        return send_file(arquivo, as_attachment=True, download_name='fonogramas.xlsx')
    
    fonogramas = fonograma_service.listar_todos_fonogramas(current_user)
    return render_template('usuario/exportar/selecao.html', fonogramas=fonogramas)

# ==================== MEU HISTÓRICO ====================

@usuario_bp.route('/historico')
@usuario_required
def meu_historico():
    """Histórico de alterações do usuário"""
    page = request.args.get('page', 1, type=int)
    
    historico = fonograma_service.obter_historico_usuario(current_user, page=page)
    
    return render_template('usuario/historico/lista.html', historico=historico)

# ==================== API JSON ====================

@usuario_bp.route('/api/fonogramas/<int:fonograma_id>/status')
@usuario_required
def api_status_fonograma(fonograma_id):
    """API: Status simplificado do fonograma"""
    fonograma = fonograma_service.obter_fonograma(fonograma_id, current_user)
    if not fonograma:
        return jsonify({'erro': 'Não encontrado'}), 404
    
    status = fonograma_service.obter_status_simplificado(fonograma)
    return jsonify(status)

@usuario_bp.route('/api/validar-campo', methods=['POST'])
@usuario_required
def api_validar_campo():
    """API: Validação de campo individual (para formulário)"""
    campo = request.json.get('campo')
    valor = request.json.get('valor')
    
    resultado = upload_service.validar_campo_individual(campo, valor)
    return jsonify(resultado)
