"""
Módulo de geração de Excel formatado no padrão ECAD/ABRAMUS
Usa layout idêntico ao template_fonograma_final.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from processador import (
    parse_autores, parse_interpretes, parse_musicos,
    parse_editoras, parse_documentos
)


# Cores institucionais SBACEM (igual ao template)
CORES_SECAO = {
    'id': PatternFill(start_color="22164C", end_color="22164C", fill_type="solid"),      # Roxo escuro
    'obra': PatternFill(start_color="3D2E6B", end_color="3D2E6B", fill_type="solid"),    # Roxo claro
    'autor': PatternFill(start_color="EF234D", end_color="EF234D", fill_type="solid"),   # Vermelho
    'interp': PatternFill(start_color="EF234D", end_color="EF234D", fill_type="solid"),  # Vermelho
    'prod': PatternFill(start_color="E1C8B0", end_color="E1C8B0", fill_type="solid"),    # Bege
    'opc': PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid"),     # Cinza
}

# Colunas na mesma ordem do template (larguras otimizadas)
COLUNAS_TEMPLATE = [
    # IDENTIFICAÇÃO (Roxo) - Campos essenciais
    {"nome": "isrc", "titulo": "ISRC", "secao": "id", "largura": 14},
    {"nome": "titulo", "titulo": "Título", "secao": "id", "largura": 22},
    {"nome": "duracao", "titulo": "Duração", "secao": "id", "largura": 9},
    {"nome": "ano_lanc", "titulo": "Ano", "secao": "id", "largura": 6},
    {"nome": "genero", "titulo": "Gênero", "secao": "id", "largura": 12},
    
    # OBRA MUSICAL (Roxo claro)
    {"nome": "titulo_obra", "titulo": "Título Obra", "secao": "obra", "largura": 22},
    
    # AUTORES (Vermelho)
    {"nome": "autores", "titulo": "Autores", "secao": "autor", "largura": 45},
    
    # INTÉRPRETES (Vermelho)
    {"nome": "interpretes", "titulo": "Intérpretes", "secao": "interp", "largura": 40},
    
    # PRODUTOR (Bege) - Todos os campos do produtor
    {"nome": "prod_nome", "titulo": "Produtor", "secao": "prod", "largura": 22},
    {"nome": "prod_doc", "titulo": "Prod. Doc", "secao": "prod", "largura": 16},
    {"nome": "prod_perc", "titulo": "Prod. %", "secao": "prod", "largura": 8},
    {"nome": "prod_fantasia", "titulo": "Prod. Fantasia", "secao": "prod", "largura": 15},
    {"nome": "prod_endereco", "titulo": "Prod. Endereço", "secao": "prod", "largura": 20},
    {"nome": "prod_assoc", "titulo": "Prod. Assoc", "secao": "prod", "largura": 12},
    {"nome": "prod_data_ini", "titulo": "Prod. Data Ini", "secao": "prod", "largura": 12},
    
    # CAMPOS OPCIONAIS (Cinza)
    {"nome": "versao", "titulo": "Versão", "secao": "opc", "largura": 10},
    {"nome": "idioma", "titulo": "Idioma", "secao": "opc", "largura": 7},
    {"nome": "ano_grav", "titulo": "Ano Grav", "secao": "opc", "largura": 8},
    {"nome": "cod_interno", "titulo": "Cód.Interno", "secao": "opc", "largura": 12},
    {"nome": "cod_obra", "titulo": "Cód.Obra", "secao": "opc", "largura": 12},
    {"nome": "editoras", "titulo": "Editoras", "secao": "opc", "largura": 30},
    {"nome": "musicos", "titulo": "Músicos", "secao": "opc", "largura": 35},
    
    # LANÇAMENTO (Cinza)
    {"nome": "tipo_lanc", "titulo": "Tipo Lanç", "secao": "opc", "largura": 10},
    {"nome": "album", "titulo": "Álbum", "secao": "opc", "largura": 18},
    {"nome": "faixa", "titulo": "Faixa", "secao": "opc", "largura": 6},
    {"nome": "selo", "titulo": "Selo", "secao": "opc", "largura": 15},
    {"nome": "formato", "titulo": "Formato", "secao": "opc", "largura": 10},
    {"nome": "pais", "titulo": "País", "secao": "opc", "largura": 12},
    {"nome": "data_lanc", "titulo": "Data Lanç", "secao": "opc", "largura": 12},
    
    # ADMINISTRATIVO (Cinza)
    {"nome": "assoc_gestao", "titulo": "Assoc. Gestão", "secao": "opc", "largura": 14},
    {"nome": "data_cad", "titulo": "Data Cad", "secao": "opc", "largura": 12},
    {"nome": "situacao", "titulo": "Situação", "secao": "opc", "largura": 10},
    {"nome": "obs_juridicas", "titulo": "Obs. Jurídicas", "secao": "opc", "largura": 20},
    {"nome": "historico", "titulo": "Histórico", "secao": "opc", "largura": 20},
    {"nome": "documentos", "titulo": "Documentos", "secao": "opc", "largura": 25},
    
    # DISTRIBUIÇÃO (Cinza)
    {"nome": "territorio", "titulo": "Território", "secao": "opc", "largura": 12},
    {"nome": "tipos_exec", "titulo": "Tipos Exec", "secao": "opc", "largura": 12},
    {"nome": "prioridade", "titulo": "Prioridade", "secao": "opc", "largura": 12},
    {"nome": "cod_ecad", "titulo": "Cód. ECAD", "secao": "opc", "largura": 14},
]


def gerar_excel(df: pd.DataFrame, caminho_saida: str):
    """
    Gera arquivo Excel formatado a partir do DataFrame processado
    Usa layout idêntico ao template_fonograma_final.xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Fonogramas"
    
    # Fontes
    font_branco = Font(bold=True, color="FFFFFF", size=10)
    font_preto = Font(bold=True, color="22164C", size=10)
    font_normal = Font(size=10)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cores para linhas alternadas
    row_fill_light = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    row_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Criar cabeçalhos
    for col_idx, col_info in enumerate(COLUNAS_TEMPLATE, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_info["titulo"])
        cell.fill = CORES_SECAO[col_info["secao"]]
        
        # Fonte branca para cores escuras, preta para bege
        if col_info["secao"] == "prod":
            cell.font = font_preto
        else:
            cell.font = font_branco
        
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_info["largura"]
    
    # Altura do cabeçalho (compacto)
    ws.row_dimensions[1].height = 25
    
    # Preencher dados
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        row_fill = row_fill_light if row_idx % 2 == 0 else row_fill_white
        
        for col_idx, col_info in enumerate(COLUNAS_TEMPLATE, 1):
            campo = col_info["nome"]
            valor = row.get(campo, '')
            
            # Valor formatado (truncar para evitar células gigantes)
            if pd.isna(valor) or valor == 'nan':
                valor_formatado = ''
            else:
                valor_str = str(valor)
                # Truncar valores muito longos (máximo 100 caracteres)
                if len(valor_str) > 100:
                    valor_formatado = valor_str[:97] + '...'
                else:
                    valor_formatado = valor_str
            
            cell = ws.cell(row=row_idx, column=col_idx, value=valor_formatado)
            cell.fill = row_fill
            cell.font = font_normal
            cell.alignment = Alignment(vertical='center', wrap_text=False)  # Sem wrap para células compactas
            cell.border = border
    
    # Congelar primeira linha e primeiras 2 colunas (ISRC e Título)
    ws.freeze_panes = 'C2'
    
    # Adicionar filtros
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    
    # Salvar
    wb.save(caminho_saida)
    return caminho_saida
